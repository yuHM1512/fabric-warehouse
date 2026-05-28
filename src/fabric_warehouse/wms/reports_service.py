from __future__ import annotations

from io import BytesIO
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import Session

from fabric_warehouse.db.models.hanging_tag import HangingTag
from fabric_warehouse.db.models.location_assignment import LocationAssignment
from fabric_warehouse.db.models.receipt import ReceiptLine
from fabric_warehouse.db.models.return_event import ReturnEvent
from fabric_warehouse.db.models.stock_check import StockCheck

_DANG_LUU = "Đang lưu"
_DANG_LUU_VARIANTS = ("Đang lưu", "Dang luu", "Đang luu", "Dang lưu")


@dataclass
class TonKhoRow:
    nhom: str
    nhom_phu: str | None
    so_cay: int
    tong_yds: float
    da_dinh_danh: int


@dataclass(frozen=True)
class AgeSplitKpis:
    under_rolls: int
    under_yds: float
    over_rolls: int
    over_yds: float

    @property
    def total_rolls(self) -> int:
        return int(self.under_rolls + self.over_rolls)

    @property
    def total_yds(self) -> float:
        return float(self.under_yds + self.over_yds)


@dataclass(frozen=True)
class StockAgeRow:
    nhu_cau: str
    lot: str
    ma_cay: str
    so_luong: float | None
    thuc_te: float | None
    ghi_chu: str | None
    vi_tri: str | None
    trang_thai: str | None
    ngay_cap_nhat: datetime | None
    assigned_at: datetime | None
    age_days: int | None
    bucket: str  # "under_6m" | "over_6m" | "unknown"


def _yds(sc: type) -> object:
    return func.coalesce(sc.actual_yards, sc.expected_yards)


def _latest_return_created_at_map(
    db: Session,
    *,
    ma_cays: list[str],
) -> dict[str, datetime]:
    if not ma_cays:
        return {}

    rows = (
        db.query(ReturnEvent.ma_cay, ReturnEvent.created_at)
        .filter(ReturnEvent.ma_cay.in_(ma_cays))
        .filter(ReturnEvent.created_at.isnot(None))
        .order_by(ReturnEvent.ma_cay.asc(), ReturnEvent.created_at.desc(), ReturnEvent.id.desc())
        .all()
    )
    out: dict[str, datetime] = {}
    for ma_cay, created_at in rows:
        ma = str(ma_cay or "").strip()
        if not ma or ma in out or created_at is None:
            continue
        out[ma] = created_at
    return out


def ton_kho_by_nhu_cau(db: Session) -> list[TonKhoRow]:
    rows = (
        db.query(
            LocationAssignment.nhu_cau.label("nhom"),
            func.count(LocationAssignment.ma_cay).label("so_cay"),
            func.coalesce(func.sum(_yds(StockCheck)), 0).label("tong_yds"),
            func.count(LocationAssignment.vi_tri).label("da_dinh_danh"),
        )
        .outerjoin(StockCheck, StockCheck.ma_cay == LocationAssignment.ma_cay)
        .filter(LocationAssignment.trang_thai == _DANG_LUU)
        .group_by(LocationAssignment.nhu_cau)
        .order_by(LocationAssignment.nhu_cau)
        .all()
    )
    return [
        TonKhoRow(
            nhom=r.nhom or "(Không xác định)",
            nhom_phu=None,
            so_cay=r.so_cay,
            tong_yds=float(r.tong_yds or 0),
            da_dinh_danh=r.da_dinh_danh,
        )
        for r in rows
    ]


def ton_kho_by_lot(db: Session) -> list[TonKhoRow]:
    rows = (
        db.query(
            LocationAssignment.lot.label("nhom"),
            LocationAssignment.nhu_cau.label("nhom_phu"),
            func.count(LocationAssignment.ma_cay).label("so_cay"),
            func.coalesce(func.sum(_yds(StockCheck)), 0).label("tong_yds"),
            func.count(LocationAssignment.vi_tri).label("da_dinh_danh"),
        )
        .outerjoin(StockCheck, StockCheck.ma_cay == LocationAssignment.ma_cay)
        .filter(LocationAssignment.trang_thai == _DANG_LUU)
        .group_by(LocationAssignment.lot, LocationAssignment.nhu_cau)
        .order_by(LocationAssignment.nhu_cau, LocationAssignment.lot)
        .all()
    )
    return [
        TonKhoRow(
            nhom=r.nhom or "(Không xác định)",
            nhom_phu=r.nhom_phu or None,
            so_cay=r.so_cay,
            tong_yds=float(r.tong_yds or 0),
            da_dinh_danh=r.da_dinh_danh,
        )
        for r in rows
    ]


def _ht_subquery(db: Session, *extra_cols):
    """One hanging_tag row per (nhu_cau, lot) — lowest id wins."""
    return (
        db.query(HangingTag.nhu_cau, HangingTag.lot, *extra_cols)
        .distinct(HangingTag.nhu_cau, HangingTag.lot)
        .order_by(HangingTag.nhu_cau, HangingTag.lot, HangingTag.id)
        .subquery()
    )


def ton_kho_by_loai_vai(db: Session) -> list[TonKhoRow]:
    ht = _ht_subquery(db, HangingTag.loai_vai)
    rows = (
        db.query(
            func.coalesce(ht.c.loai_vai, "(Không xác định)").label("nhom"),
            func.count(LocationAssignment.ma_cay).label("so_cay"),
            func.coalesce(func.sum(_yds(StockCheck)), 0).label("tong_yds"),
            func.count(LocationAssignment.vi_tri).label("da_dinh_danh"),
        )
        .outerjoin(ht, (ht.c.nhu_cau == LocationAssignment.nhu_cau) & (ht.c.lot == LocationAssignment.lot))
        .outerjoin(StockCheck, StockCheck.ma_cay == LocationAssignment.ma_cay)
        .filter(LocationAssignment.trang_thai == _DANG_LUU)
        .group_by(ht.c.loai_vai)
        .order_by(ht.c.loai_vai)
        .all()
    )
    return [
        TonKhoRow(
            nhom=r.nhom or "(Không xác định)",
            nhom_phu=None,
            so_cay=r.so_cay,
            tong_yds=float(r.tong_yds or 0),
            da_dinh_danh=r.da_dinh_danh,
        )
        for r in rows
    ]


def ton_kho_by_mau_vai(db: Session) -> list[TonKhoRow]:
    ht = _ht_subquery(db, HangingTag.mau_vai, HangingTag.ma_mau)
    rows = (
        db.query(
            func.coalesce(ht.c.mau_vai, "(Không xác định)").label("nhom"),
            func.coalesce(ht.c.ma_mau, "").label("nhom_phu"),
            func.count(LocationAssignment.ma_cay).label("so_cay"),
            func.coalesce(func.sum(_yds(StockCheck)), 0).label("tong_yds"),
            func.count(LocationAssignment.vi_tri).label("da_dinh_danh"),
        )
        .outerjoin(ht, (ht.c.nhu_cau == LocationAssignment.nhu_cau) & (ht.c.lot == LocationAssignment.lot))
        .outerjoin(StockCheck, StockCheck.ma_cay == LocationAssignment.ma_cay)
        .filter(LocationAssignment.trang_thai == _DANG_LUU)
        .group_by(ht.c.mau_vai, ht.c.ma_mau)
        .order_by(ht.c.mau_vai)
        .all()
    )
    return [
        TonKhoRow(
            nhom=r.nhom or "(Không xác định)",
            nhom_phu=r.nhom_phu or None,
            so_cay=r.so_cay,
            tong_yds=float(r.tong_yds or 0),
            da_dinh_danh=r.da_dinh_danh,
        )
        for r in rows
    ]


def ton_kho_by_age_split(
    db: Session,
    *,
    limit: int = 5000,
    split_days: int = 183,
    bucket: str | None = None,  # "under_6m" | "over_6m" | None
    nhu_cau: str | None = None,
    lot: str | None = None,
    sort: str = "nearest",  # "nearest" | "farthest"
) -> tuple[AgeSplitKpis, list[StockAgeRow]]:
    """
    Rolls currently stored, split into 2 buckets:
    - under_6m: assigned_at >= now - split_days
    - over_6m:  assigned_at <  now - split_days

    Uses LocationAssignment.assigned_at as the stored-confirmation timestamp.
    """
    now = datetime.now(timezone.utc)
    split_at = now - timedelta(days=int(split_days))

    q = (
        db.query(LocationAssignment, StockCheck)
        .outerjoin(
            StockCheck,
            and_(
                StockCheck.ma_cay == LocationAssignment.ma_cay,
                StockCheck.nhu_cau == LocationAssignment.nhu_cau,
                StockCheck.lot == LocationAssignment.lot,
            ),
        )
        .filter(LocationAssignment.trang_thai == _DANG_LUU)
    )

    if nhu_cau:
        q = q.filter(LocationAssignment.nhu_cau == nhu_cau)
    if lot:
        q = q.filter(LocationAssignment.lot == lot)

    if bucket == "under_6m":
        q = q.filter(LocationAssignment.assigned_at.isnot(None)).filter(LocationAssignment.assigned_at >= split_at)
    elif bucket == "over_6m":
        q = q.filter(LocationAssignment.assigned_at.isnot(None)).filter(LocationAssignment.assigned_at < split_at)

    if sort == "farthest":
        q = q.order_by(LocationAssignment.assigned_at.asc().nulls_last(), LocationAssignment.updated_at.desc())
    else:
        q = q.order_by(LocationAssignment.assigned_at.desc().nulls_last(), LocationAssignment.updated_at.desc())

    pairs = q.limit(limit).all()
    ma_cays = [str(la.ma_cay or "").strip() for la, _ in pairs if getattr(la, "ma_cay", None)]
    return_created_at_by_ma_cay = _latest_return_created_at_map(db, ma_cays=ma_cays)

    under_rolls = 0
    over_rolls = 0
    under_yds = 0.0
    over_yds = 0.0

    out: list[StockAgeRow] = []
    for la, sc in pairs:
        nhu_cau = str(la.nhu_cau or "").strip()
        lot = str(la.lot or "").strip()
        ma_cay = str(la.ma_cay or "").strip()

        so_luong = float(sc.expected_yards) if (sc and sc.expected_yards is not None) else None
        thuc_te = float(sc.actual_yards) if (sc and sc.actual_yards is not None) else None
        ghi_chu = (sc.note if sc else None) or None

        vi_tri = str(la.vi_tri).strip() if la.vi_tri is not None else None
        trang_thai = str(la.trang_thai).strip() if la.trang_thai is not None else None

        assigned_at = getattr(la, "assigned_at", None)
        updated_at = getattr(la, "updated_at", None) or (getattr(sc, "updated_at", None) if sc else None)
        returned_at = return_created_at_by_ma_cay.get(ma_cay)
        if returned_at and (updated_at is None or returned_at > updated_at):
            updated_at = returned_at

        age_days: int | None = None
        bucket = "unknown"
        yds_val = (thuc_te if thuc_te is not None else so_luong) or 0.0

        if isinstance(assigned_at, datetime):
            try:
                age_days = int((now - assigned_at.astimezone(timezone.utc)).days)
            except Exception:
                age_days = None

            if assigned_at < split_at:
                bucket = "over_6m"
                over_rolls += 1
                over_yds += float(yds_val)
            else:
                bucket = "under_6m"
                under_rolls += 1
                under_yds += float(yds_val)

        out.append(
            StockAgeRow(
                nhu_cau=nhu_cau,
                lot=lot,
                ma_cay=ma_cay,
                so_luong=so_luong,
                thuc_te=thuc_te,
                ghi_chu=ghi_chu,
                vi_tri=vi_tri,
                trang_thai=trang_thai,
                ngay_cap_nhat=updated_at,
                assigned_at=assigned_at,
                age_days=age_days,
                bucket=bucket,
            )
        )

    return (
        AgeSplitKpis(
            under_rolls=int(under_rolls),
            under_yds=float(under_yds),
            over_rolls=int(over_rolls),
            over_yds=float(over_yds),
        ),
        out,
    )


@dataclass(frozen=True)
class InboundLotRow:
    nhu_cau: str
    lot: str
    vi_tri: str | None
    so_cay: int
    da_nhap_cay: int
    tong_yds: float
    da_nhap_yds: float
    is_complete: bool


@dataclass(frozen=True)
class InboundDemandGroup:
    nhu_cau: str
    lots: list[InboundLotRow]
    status: str

    @property
    def rowspan(self) -> int:
        return len(self.lots)

    @property
    def lot_count(self) -> int:
        return len({row.lot for row in self.lots})


@dataclass(frozen=True)
class StockExportRow:
    nhu_cau: str
    lot: str
    vi_tri: str | None
    ma_cay: str
    so_yds: float

def inbound_status_by_nhu_cau(
    db: Session,
    *,
    nhu_cau: str | None = None,
    status: str | None = None,
    limit_lots: int = 5000,
) -> list[InboundDemandGroup]:
    rr = (
        db.query(
            ReceiptLine.nhu_cau.label("nhu_cau"),
            ReceiptLine.lot.label("lot"),
            ReceiptLine.ma_cay.label("ma_cay"),
            func.max(ReceiptLine.id).label("rid"),
        )
        .filter(ReceiptLine.nhu_cau.isnot(None))
        .filter(ReceiptLine.lot.isnot(None))
        .group_by(ReceiptLine.nhu_cau, ReceiptLine.lot, ReceiptLine.ma_cay)
        .subquery()
    )
    rl = db.query(ReceiptLine).subquery()

    q = (
        db.query(
            rr.c.nhu_cau,
            rr.c.lot,
            LocationAssignment.vi_tri.label("vi_tri"),
            func.count(rr.c.ma_cay).label("so_cay"),
            func.count(func.distinct(StockCheck.ma_cay)).label("da_nhap_cay"),
            func.coalesce(func.sum(func.coalesce(rl.c.yards, 0)), 0).label("tong_yds"),
            func.coalesce(
                func.sum(func.coalesce(StockCheck.actual_yards, StockCheck.expected_yards, 0)),
                0,
            ).label("da_nhap_yds"),
        )
        .join(rl, rl.c.id == rr.c.rid)
        .outerjoin(
            StockCheck,
            and_(
                StockCheck.ma_cay == rr.c.ma_cay,
                StockCheck.nhu_cau == rr.c.nhu_cau,
                StockCheck.lot == rr.c.lot,
            ),
        )
        .outerjoin(
            LocationAssignment,
            and_(
                LocationAssignment.ma_cay == rr.c.ma_cay,
                LocationAssignment.nhu_cau == rr.c.nhu_cau,
                LocationAssignment.lot == rr.c.lot,
            ),
        )
    )
    if nhu_cau:
        q = q.filter(rr.c.nhu_cau.ilike(f"%{nhu_cau}%"))

    rows = (
        q.group_by(rr.c.nhu_cau, rr.c.lot, LocationAssignment.vi_tri)
        .order_by(rr.c.nhu_cau, rr.c.lot, LocationAssignment.vi_tri.nullsfirst())
        .limit(int(limit_lots))
        .all()
    )

    grouped: dict[str, dict[str, object]] = {}
    for nc, lt, vi_tri, so_cay, da_nhap_cay, tong_yds, da_nhap_yds in rows:
        nc_s = str(nc or "").strip() or "(Khong xac dinh)"
        lt_s = str(lt or "").strip() or "(Khong xac dinh)"
        vi_tri_s = str(vi_tri).strip() if vi_tri else None
        so_cay_i = int(so_cay or 0)
        da_nhap_cay_i = int(da_nhap_cay or 0)
        tong_yds_f = float(tong_yds or 0)
        da_nhap_yds_f = float(da_nhap_yds or 0)
        lot_complete = so_cay_i > 0 and da_nhap_cay_i >= so_cay_i and da_nhap_yds_f + 1e-6 >= tong_yds_f
        payload = grouped.setdefault(
            nc_s,
            {
                "lots": [],
                "total_rolls": 0,
                "total_yds": 0.0,
                "checked_rolls": 0,
                "checked_yds": 0.0,
            },
        )
        payload["lots"].append(
            InboundLotRow(
                nhu_cau=nc_s,
                lot=lt_s,
                vi_tri=vi_tri_s,
                so_cay=so_cay_i,
                da_nhap_cay=da_nhap_cay_i,
                tong_yds=tong_yds_f,
                da_nhap_yds=da_nhap_yds_f,
                is_complete=lot_complete,
            )
        )
        payload["total_rolls"] = int(payload["total_rolls"]) + so_cay_i
        payload["total_yds"] = float(payload["total_yds"]) + tong_yds_f
        payload["checked_rolls"] = int(payload["checked_rolls"]) + da_nhap_cay_i
        payload["checked_yds"] = float(payload["checked_yds"]) + da_nhap_yds_f

    groups: list[InboundDemandGroup] = []
    for nc, payload in grouped.items():
        total_rolls = int(payload["total_rolls"])
        total_yds = float(payload["total_yds"])
        checked_rolls = int(payload["checked_rolls"])
        checked_yds = float(payload["checked_yds"])
        demand_status = (
            "nhap_du"
            if total_rolls > 0 and checked_rolls >= total_rolls and checked_yds + 1e-6 >= total_yds
            else "dang_nhap_kho"
        )
        if status and status != demand_status:
            continue
        groups.append(
            InboundDemandGroup(
                nhu_cau=nc,
                lots=list(payload["lots"]),
                status=demand_status,
            )
        )

    return groups


def list_active_inbound_nhu_cau_options(
    db: Session,
    *,
    status: str | None = None,
    limit: int = 5000,
) -> list[str]:
    groups = inbound_status_by_nhu_cau(db, status=status, limit_lots=limit)
    return [g.nhu_cau for g in groups[:limit]]


def list_stock_export_rows(
    db: Session,
    *,
    nhu_cau: str | None = None,
) -> list[StockExportRow]:
    q = (
        db.query(
            ReceiptLine.nhu_cau.label("nhu_cau"),
            ReceiptLine.lot.label("lot"),
            LocationAssignment.vi_tri.label("vi_tri"),
            ReceiptLine.ma_cay.label("ma_cay"),
            func.coalesce(ReceiptLine.yards, 0).label("so_yds"),
        )
        .join(LocationAssignment, LocationAssignment.ma_cay == ReceiptLine.ma_cay)
        .filter(LocationAssignment.trang_thai == _DANG_LUU)
    )
    if nhu_cau:
        q = q.filter(ReceiptLine.nhu_cau.ilike(f"%{nhu_cau}%"))

    rows = (
        q.order_by(
            ReceiptLine.nhu_cau.asc(),
            ReceiptLine.lot.asc(),
            LocationAssignment.vi_tri.asc(),
            ReceiptLine.ma_cay.asc(),
        ).all()
    )
    return [
        StockExportRow(
            nhu_cau=str(r.nhu_cau or "").strip() or "(Khong xac dinh)",
            lot=str(r.lot or "").strip() or "(Khong xac dinh)",
            vi_tri=str(r.vi_tri).strip() if r.vi_tri else None,
            ma_cay=str(r.ma_cay or "").strip(),
            so_yds=float(r.so_yds or 0),
        )
        for r in rows
    ]


def build_stock_export_excel(
    db: Session,
    *,
    nhu_cau: str | None = None,
    export_mode: str = "selected",
) -> bytes:
    rows = list_stock_export_rows(db, nhu_cau=nhu_cau if export_mode == "selected" else None)

    wb = Workbook()
    ws = wb.active
    ws.title = "TonKho"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    title = "BÁO CÁO TỒN KHO ĐANG LƯU"
    subtitle = f"Nhu cầu: {nhu_cau}" if export_mode == "selected" and nhu_cau else "Phạm vi: ALL đang lưu kho"
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws.merge_cells("A1:E1")
    ws["A1"] = title
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:E2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=11, bold=True, color="1F1F1F")
    ws["A2"].fill = PatternFill("solid", fgColor="D9EAF7")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A3:E3")
    ws["A3"] = f"Generated at: {generated_at}"
    ws["A3"].font = Font(size=10, italic=True, color="666666")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Nhu cầu", "Lot", "Vị trí", "Mã cây", "Số YDS"]
    header_row = 4
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F75B5")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_yds = 0.0
    start_data_row = header_row + 1
    for offset, item in enumerate(rows):
        row_idx = start_data_row + offset
        values = [item.nhu_cau, item.lot, item.vi_tri or "", item.ma_cay, item.so_yds]
        total_yds += item.so_yds
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if col_idx == 5:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        if offset % 2 == 0:
            for col_idx in range(1, 6):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor="F7FBFF")

    summary_row = max(start_data_row, start_data_row + len(rows))
    ws.cell(row=summary_row, column=1, value="Tổng cộng").font = Font(bold=True)
    ws.cell(row=summary_row, column=1).fill = PatternFill("solid", fgColor="EAF2F8")
    ws.cell(row=summary_row, column=1).border = border
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
    for col_idx in range(1, 5):
        ws.cell(row=summary_row, column=col_idx).border = border
        ws.cell(row=summary_row, column=col_idx).fill = PatternFill("solid", fgColor="EAF2F8")
    total_cell = ws.cell(row=summary_row, column=5, value=total_yds)
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00'
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.fill = PatternFill("solid", fgColor="EAF2F8")
    total_cell.border = border

    ws.auto_filter.ref = f"A4:E{summary_row}"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[4].height = 22

    widths = {
        "A": 24,
        "B": 18,
        "C": 16,
        "D": 22,
        "E": 14,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=4, max_row=summary_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = border

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_title_rows = "$1:$4"

    for idx in range(1, 6):
        ws.cell(row=summary_row, column=idx).border = border
        ws.column_dimensions[get_column_letter(idx)].bestFit = True

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

